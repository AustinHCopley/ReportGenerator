import openpyxl
from pandas import read_csv

class Generator():
    """client sales report generator"""

    def __init__(self, csv='transactions.csv'): #by default, opens a csv file called transactions
        self.csv = csv
        self.xlsx = csv.split('.')[0] + '.xlsx'
        self.assets = []
        self.toXL()
        # get the initial "Total" sheet from the original csv file
        self.Total = self.wb.active
        
    def toXL(self):
        """convert csv to xlsx with pandas and open it as workbood with openpyxl"""
        try:
            self.transactions = read_csv(self.csv)
        except FileNotFoundError:
            print(f"File: {self.xlsx} not found")
            exit()
        self.transactions.to_excel(self.xlsx, sheet_name="Total", index=False)

        try:
            # create the instance's workbook
            self.wb = openpyxl.load_workbook(self.xlsx)
        except openpyxl.utils.exceptions.InvalidFileException:
            print(f"File: {self.xlsx} is not a valid excel file")
            exit()
    
    def setHeading(self, heading):
        self.heading = heading

    def setAssets(self, assets):
        self.assets = assets
    
    def addAsset(self, value):
        self.assets.append(value)

    @staticmethod
    def insertRow(sheet, row, j = 1):
        for i, col in enumerate(row):
            sheet.cell(row=j, column=i+1).value = col.value


def insertionSortA(list, key):
    """
    insertion sort algorithm, refactored to be tailored to the specific needs of the program (sorting asset tuples by different indices)
    the key is the index of the tuple item to sort the list by
    """
    for i in range(1, len(list)):
        key_item = list[i][key]
        j = i - 1
        while j >= 0 and str(list[j][key]) < str(key_item):
            list[j + 1] = list[j]
            j -= 1
        list[j + 1] = key_item
    return list

def insertionSortD(list, key):
    """
    Descending version of insertion sort
    """

    for i in range(1, len(list)):
        print(list[i], key)
        key_item = list[i][key]
        j = i - 1
        while j >= 0 and str(list[j][key]) < str(key_item):
            print("list[j]][key]: ", list[j][key], "key_item: ", key_item)
            list[j + 1] = list[j]
            j -= 1

        list[j + 1] = key_item

    return list

def main():

    # create generator instance
    gen = Generator("transactions.csv")

    # delete G-J, removing extraneous data that clients dont need to see
    gen.Total.delete_cols(7, 4)
    # save the header for copying into new sheets
    gen.setHeading(gen.Total[1])

    # Create separate sheets:
    # for each item in column B, if the assetID is new, create a new sheet
    tempAssets = []
    for cell in gen.Total['B']:
        if cell.value != "Asset ID":
            if cell.value not in tempAssets:
                tempAssets.append(cell.value)
                gen.wb.create_sheet( title=f"p{tempAssets.index(cell.value)+1}" )
                tempSheet = gen.wb[ f"p{tempAssets.index(cell.value)+1}" ]
                gen.insertRow(tempSheet, gen.heading)

        # --> Find Assets
    # list assets with (ID, name, price)
    for row in gen.Total.iter_rows(min_row=2, max_row=gen.Total.max_row, min_col=1, max_col=9):
        if ( row[1].value, row[8].value, row[4].value ) not in gen.assets and row[4].value != 0: # if (id, name, price) tuple is not in assets
            print("no", row[1].value, row[8].value, row[4].value)
            gen.addAsset( (row[1].value, row[8].value, row[4].value) ) # append a tuple with (ID, name, price) to be sorted alphabetically and then based on price high to low

    print("Assets <<<", gen.assets, ">>>")
    # sort the list of assets alphabetically by name
    print("A" > "B")
    gen.setAssets( insertionSortA(gen.assets, 1) )

    # sort the list of assets by price high to low
    #gen.setAssets( insertionSortD(gen.assets, 2) )

    print("Assets <<<", gen.assets, ">>>")
    
    # add tab for vouchers and refunds
    vouchers = gen.wb.create_sheet(title="vouchers")
    gen.insertRow(vouchers, gen.heading)
    refunds = gen.wb.create_sheet(title="refunds")
    gen.insertRow(refunds, gen.heading)
    
    # TODO: remove vouchers and refunds from total tab before finding assets
    ### --> Loop through each row
    v = 2
    r = 2
    for row in gen.Total.iter_rows(min_row=2, max_row=gen.Total.max_row, min_col=1, max_col=19):
        # separate vouchers into separate sheet
        # all transactions with a value in "Note" column
        if row[6].value is not None:
            gen.insertRow(vouchers, row, v)
            v += 1

        # separate refunds into separate sheet
        # all remaining transactions without a value in "Referrer" column
        elif row[7].value is None:
            gen.insertRow(refunds, row, r)
            r += 1
        
        else:
            #  TODO: seperate transactions according to assetID into proper sheet
            print(row[1].value) #show asset id
            #print(gen.assets.index(row[1].value)) #show index of asset id

    gen.wb.save(gen.xlsx)


if __name__ == "__main__":
    main()
    print("Process complete")
    exit()